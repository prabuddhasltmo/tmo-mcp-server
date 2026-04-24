"""
TMO Loan Aging Report  –  v2
6 vertical aging-bucket columns with ✓ checkmarks, one row per loan.
Usage: python3 aging_report_v2.py <loans_json_file>
"""
import sys, json, datetime
from pathlib import Path

# ── Load loans ────────────────────────────────────────────────────────────────
LOANS_FILE = sys.argv[1]
with open(LOANS_FILE) as f:
    raw = json.load(f)

# Support both "array of tool-call results" wrapper or plain API response
if isinstance(raw, list) and raw and isinstance(raw[0], dict) and 'text' in raw[0]:
    loans_all = json.loads(raw[0]['text'])['Data']
elif isinstance(raw, dict):
    loans_all = raw.get('Data', raw)
else:
    loans_all = raw

# ── Lender / Portfolio mapping (from LSS lender portfolio – sandbox data) ─────
LENDER_NAMES = {
    'BROKER':   'World Mortgage Co. (Broker)',
    'COMPANY':  'World Mortgage Co.',
    'LENDER-B': 'Financial Partners, LLC',
    'LENDER-C': 'Ontario Mortgage Inv. Corp.',
    'LENDER-D': 'AB Mortgage Inv. Corp.',
    'LENDER-E': 'NY Equity Investment Fund',
    'LENDER-F': 'California Capital Group',
    'LENDER-G': 'Mortgage Opportunity Fund',
    'MI10':     'Private Capital Lending, LLC',
}

PORTFOLIO_DATA = {
    'BROKER':   [('B001001',6.065)],
    'COMPANY':  [('B001001',26.683),('B001002',100),('B001003',100),('B001004',100),
                 ('B001005',25.128),('B001006',100),('B001007',100),('B001008',35.201),
                 ('B001009',100),('B001010',42.538),('B001011',29.937),('B001012',100),
                 ('B001013',100),('B001014',100),('B001015',29.126),('B001016',30.354),
                 ('B001017',100),('B001018',100),('B001019',32.525),('B001020',28.531),
                 ('B001021',100),('B001031',25),('B001032',25),('B001033',60),
                 ('B001034',33.333),('B001036',50),('B001037',50),('B001038',50),
                 ('B001039',50),('B001040',50),('B001041',50),('B001042',50),
                 ('B001043',50),('B001044',50),('B001045',50),('B001049',50),
                 ('B001050',50),('B001051',50)],
    'LENDER-B': [('B001001',17.407),('B001027',34.712),('B001034',33.333),('B001035',50),
                 ('B001036',25),('B001037',25),('B001038',25),('B001039',25),('B001040',25),
                 ('B001041',25),('B001042',25),('B001043',25),('B001044',25),('B001045',25),
                 ('B001046',25),('B001047',25),('B001048',33.333),('B001049',25),
                 ('B001050',25),('B001051',25)],
    'LENDER-C': [('B001001',32.439),('B001008',24.3),('B001016',43.529),
                 ('B001027',19.284),('B001030',100)],
    'LENDER-D': [('B001008',40.499),('B001016',26.117),('B001019',42.172),('B001029',100)],
    'LENDER-E': [('B001005',49.915),('B001010',34.477),('B001015',29.531),
                 ('B001019',25.303),('B001020',38.983)],
    'LENDER-F': [('B001005',24.957),('B001010',22.985),('B001011',31.847),('B001020',32.486)],
    'LENDER-G': [('B001011',38.216),('B001015',41.343),('B001028',100)],
    'MI10':     [('B001001',17.407),('B001022',100),('B001023',100),('B001024',100),
                 ('B001025',100),('B001026',100),('B001027',46.004),('B001031',75),
                 ('B001032',75),('B001033',40),('B001034',33.333),('B001035',50),
                 ('B001036',25),('B001037',25),('B001038',25),('B001039',25),('B001040',25),
                 ('B001041',25),('B001042',25),('B001043',25),('B001044',25),('B001045',25),
                 ('B001046',75),('B001047',75),('B001048',66.667),('B001049',25),
                 ('B001050',25),('B001051',25)],
}

# Build loan → primary lender map
loan_lender_map = {}
for lacct, entries in PORTFOLIO_DATA.items():
    for loan, pct in entries:
        if pct > 0:
            loan_lender_map.setdefault(loan, {})[lacct] = pct

def primary_lender(acct):
    lmap = loan_lender_map.get(acct, {})
    if not lmap:
        return 'Unassigned'
    top = max(lmap, key=lmap.get)
    return LENDER_NAMES.get(top, top)

# ── Aging buckets ──────────────────────────────────────────────────────────────
BUCKETS = [
    ('Current',      0,   0),
    ('1-30 Days',    1,  30),
    ('31-60 Days',  31,  60),
    ('61-90 Days',  61,  90),
    ('91-120 Days', 91, 120),
    ('121+ Days',  121, 99999),
]
BUCKET_LABELS = [b[0] for b in BUCKETS]

def get_bucket(days):
    for label, lo, hi in BUCKETS:
        if lo <= days <= hi:
            return label
    return '121+ Days'

# ── Build rows ─────────────────────────────────────────────────────────────────
rows = []
for loan in loans_all:
    bal = float(loan.get('PrincipalBalance') or 0)
    if bal <= 0:
        continue
    days   = int(loan.get('DaysLate') or 0)
    acct   = loan.get('Account', '')
    bname  = loan.get('SortName', '') or loan.get('ByLastName', '') or ''
    if isinstance(bname, dict):
        bname = ''
    rows.append({
        'bucket':   get_bucket(days),
        'account':  acct,
        'borrower': bname.strip(),
        'lender':   primary_lender(acct),
        'balance':  bal,
        'days':     days,
    })

rows.sort(key=lambda r: (-r['balance']))

total_bal   = sum(r['balance'] for r in rows)
bucket_data = {
    lbl: [r for r in rows if r['bucket'] == lbl]
    for lbl in BUCKET_LABELS
}

# ── Excel ──────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed – run: pip3 install openpyxl")
    sys.exit(1)

today_str = datetime.date.today().strftime('%B %-d, %Y')

BUCKET_COLORS = {
    'Current':     ('00B050', 'FFFFFF'),   # green
    '1-30 Days':   ('FFFF00', '000000'),   # yellow
    '31-60 Days':  ('FFC000', '000000'),   # orange
    '61-90 Days':  ('FF0000', 'FFFFFF'),   # red
    '91-120 Days': ('C00000', 'FFFFFF'),   # dark red
    '121+ Days':   ('7030A0', 'FFFFFF'),   # purple
}

wb = openpyxl.Workbook()

thin   = Side(style='thin',   color='BFBFBF')
thick  = Side(style='medium', color='888888')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
right  = Alignment(horizontal='right',  vertical='center')
left   = Alignment(horizontal='left',   vertical='center')

title_fill   = PatternFill('solid', fgColor='1F3864')
header_fill  = PatternFill('solid', fgColor='2E75B6')
total_fill   = PatternFill('solid', fgColor='D6E4F0')
alt_fill     = PatternFill('solid', fgColor='F2F7FC')

title_font  = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
hdr_font    = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
body_font   = Font(name='Calibri', size=10)
bold_font   = Font(name='Calibri', bold=True, size=10)
total_font  = Font(name='Calibri', bold=True, size=11)
check_font  = Font(name='Calibri', bold=True, size=12, color='000000')

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 1 – Checkmark Grid
# Columns: Lender | Loan | Primary Borrower | Principal Balance |
#           Current | 1-30 | 31-60 | 61-90 | 91-120 | 121+
# ════════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Aging Detail'

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells('A1:J1')
ws['A1'] = f'TMO Loan Aging Report  —  {today_str}'
ws['A1'].font      = title_font
ws['A1'].fill      = title_fill
ws['A1'].alignment = center
ws.row_dimensions[1].height = 30

# ── Sub-header: bucket totals (row 2) ────────────────────────────────────────
ws.merge_cells('A2:D2')
ws['A2'] = f'{len(rows)} Active Loans  |  Total Portfolio: ${total_bal:,.2f}'
ws['A2'].font      = Font(name='Calibri', bold=True, size=10, color='1F3864')
ws['A2'].alignment = left

for bi, lbl in enumerate(BUCKET_LABELS):
    col = 5 + bi   # columns E–J
    cnt = len(bucket_data[lbl])
    bal = sum(r['balance'] for r in bucket_data[lbl])
    bg, fg = BUCKET_COLORS[lbl]
    c = ws.cell(row=2, column=col,
                value=f'{cnt} loans\n${bal:,.0f}')
    c.fill      = PatternFill('solid', fgColor=bg)
    c.font      = Font(name='Calibri', bold=True, size=9, color=fg)
    c.alignment = center
ws.row_dimensions[2].height = 28

# ── Column headers (row 3) ────────────────────────────────────────────────────
FIXED_HEADERS = ['Primary Lender', 'Loan Account', 'Primary Borrower', 'Principal Balance']
BUCKET_HEADERS = BUCKET_LABELS  # 6 bucket columns

for ci, hdr in enumerate(FIXED_HEADERS, 1):
    c = ws.cell(row=3, column=ci, value=hdr)
    c.font      = hdr_font
    c.fill      = header_fill
    c.border    = border
    c.alignment = center

for bi, lbl in enumerate(BUCKET_LABELS):
    col = 5 + bi
    bg, fg = BUCKET_COLORS[lbl]
    c = ws.cell(row=3, column=col, value=lbl)
    c.font      = Font(name='Calibri', bold=True, size=10, color=fg)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.border    = border
    c.alignment = center

ws.row_dimensions[3].height = 22

# ── Data rows ─────────────────────────────────────────────────────────────────
for ri, row in enumerate(rows):
    excel_row = 4 + ri
    fill = alt_fill if ri % 2 == 1 else None

    # Fixed columns
    vals = [row['lender'], row['account'], row['borrower'], row['balance']]
    aligns = [left, center, left, right]
    for ci, (val, aln) in enumerate(zip(vals, aligns), 1):
        c = ws.cell(row=excel_row, column=ci, value=val)
        c.font      = body_font
        c.border    = border
        c.alignment = aln
        if fill:
            c.fill = fill
    ws.cell(row=excel_row, column=4).number_format = '$#,##0.00'

    # 6 bucket checkmark columns
    for bi, lbl in enumerate(BUCKET_LABELS):
        col  = 5 + bi
        bg, fg = BUCKET_COLORS[lbl]
        is_match = (row['bucket'] == lbl)
        c = ws.cell(row=excel_row, column=col,
                    value='✓' if is_match else '')
        c.border    = border
        c.alignment = center
        if is_match:
            c.font = Font(name='Calibri', bold=True, size=13, color=fg)
            c.fill = PatternFill('solid', fgColor=bg)
        else:
            c.font = body_font
            if fill:
                c.fill = fill

# ── Totals row ────────────────────────────────────────────────────────────────
total_row = 4 + len(rows)
ws.merge_cells(f'A{total_row}:C{total_row}')
c = ws.cell(row=total_row, column=1, value=f'TOTAL  —  {len(rows)} Loans')
c.font = total_font; c.fill = total_fill; c.border = border; c.alignment = right

c = ws.cell(row=total_row, column=4, value=total_bal)
c.font = total_font; c.fill = total_fill; c.border = border
c.alignment = right; c.number_format = '$#,##0.00'

for bi, lbl in enumerate(BUCKET_LABELS):
    col = 5 + bi
    cnt = len(bucket_data[lbl])
    bg, fg = BUCKET_COLORS[lbl]
    c = ws.cell(row=total_row, column=col, value=cnt)
    c.font      = Font(name='Calibri', bold=True, size=10, color=fg)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.border    = border
    c.alignment = center

ws.row_dimensions[total_row].height = 18

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 32   # Lender
ws.column_dimensions['B'].width = 14   # Loan Account
ws.column_dimensions['C'].width = 30   # Borrower
ws.column_dimensions['D'].width = 18   # Balance
ws.column_dimensions['E'].width = 10   # Current
ws.column_dimensions['F'].width = 10   # 1-30
ws.column_dimensions['G'].width = 10   # 31-60
ws.column_dimensions['H'].width = 10   # 61-90
ws.column_dimensions['I'].width = 11   # 91-120
ws.column_dimensions['J'].width = 10   # 121+

# Freeze panes so header stays visible
ws.freeze_panes = 'A4'

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 2 – Summary
# ════════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Summary')

ws2.merge_cells('A1:F1')
ws2['A1'] = f'TMO Loan Aging Report  —  Summary  |  {today_str}'
ws2['A1'].font      = title_font
ws2['A1'].fill      = title_fill
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 30

# Summary headers
sum_hdrs = ['Aging Bucket', '# Loans', 'Portfolio Balance', '% of Total', 'Avg Days Late', 'Avg Loan Size']
for ci, h in enumerate(sum_hdrs, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = hdr_font; c.fill = header_fill
    c.border = border; c.alignment = center
ws2.row_dimensions[2].height = 20

for bi, (lbl, _, __) in enumerate(BUCKETS):
    br  = bucket_data[lbl]
    cnt = len(br)
    bal = sum(r['balance'] for r in br)
    pct = (bal / total_bal) if total_bal else 0
    avg_days = (sum(r['days'] for r in br) / cnt) if cnt else 0
    avg_size = (bal / cnt) if cnt else 0
    bg, fg   = BUCKET_COLORS[lbl]
    row_i    = 3 + bi
    for ci, val in enumerate([lbl, cnt, bal, pct, avg_days, avg_size], 1):
        c = ws2.cell(row=row_i, column=ci, value=val)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.font      = Font(name='Calibri', size=10, color=fg, bold=(ci==1))
        c.border    = border
        c.alignment = center if ci in (1,2) else right
    ws2.cell(row=row_i, column=3).number_format = '$#,##0.00'
    ws2.cell(row=row_i, column=4).number_format = '0.0%'
    ws2.cell(row=row_i, column=5).number_format = '0.0'
    ws2.cell(row=row_i, column=6).number_format = '$#,##0.00'

# Summary totals
tr = 3 + len(BUCKETS)
avg_days_all = (sum(r['days'] for r in rows)/len(rows)) if rows else 0
avg_size_all = (total_bal/len(rows)) if rows else 0
for ci, val in enumerate(['TOTAL', len(rows), total_bal, 1.0, avg_days_all, avg_size_all], 1):
    c = ws2.cell(row=tr, column=ci, value=val)
    c.font = total_font; c.fill = total_fill
    c.border = border; c.alignment = center if ci in (1,2) else right
ws2.cell(row=tr, column=3).number_format = '$#,##0.00'
ws2.cell(row=tr, column=4).number_format = '0.0%'
ws2.cell(row=tr, column=5).number_format = '0.0'
ws2.cell(row=tr, column=6).number_format = '$#,##0.00'

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 18

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = Path('/Users/prabuddhalakshminarayana/Documents/TMO Repos/TMO_Loan_Aging_Report.xlsx')
wb.save(out_path)
print(f'\n✅  Excel saved → {out_path}')
print(f'   {len(rows)} loans  |  Total: ${total_bal:,.2f}')
for lbl in BUCKET_LABELS:
    cnt = len(bucket_data[lbl])
    bal = sum(r['balance'] for r in bucket_data[lbl])
    print(f'   {lbl:<14}  {cnt:>3} loans  ${bal:>18,.2f}')
