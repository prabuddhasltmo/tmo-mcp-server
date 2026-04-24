import sys, json, os

LOANS_FILE = sys.argv[1]

with open(LOANS_FILE) as f:
    raw = f.read()

loans = json.loads(json.loads(raw)[0]['text'])['Data']

lender_names = {
    'BROKER':   'World Mortgage Co. (Broker)',
    'COMPANY':  'World Mortgage Co.',
    'LENDER-B': 'Financial Partners, LLC',
    'LENDER-C': 'Ontario Mortgage Inv. Corp.',
    'LENDER-D': 'AB Mortgage Inv. Corp.',
    'LENDER-E': 'NY Equity Investment Fund',
    'LENDER-F': 'California Capital Group',
    'LENDER-G': 'Mortgage Opportunity Fund',
    'MI10':     'Private Capital Lending, LLC',
}

portfolio_data = {
    'BROKER':   [('B001001',6.065)],
    'COMPANY':  [('B001001',26.683),('B001002',100),('B001003',100),('B001004',100),
                 ('B001005',25.128),('B001006',100),('B001007',100),('B001008',35.201),
                 ('B001009',100),('B001010',42.538),('B001011',29.937),('B001012',100),
                 ('B001013',100),('B001014',100),('B001015',29.126),('B001016',30.354),
                 ('B001017',100),('B001018',100),('B001019',32.525),('B001020',28.531),
                 ('B001021',100),('B001031',25),('B001032',25),('B001033',60),
                 ('B001034',33.333),('B001036',50),('B001037',50),('B001038',50),
                 ('B001039',50),('B001040',50),('B001041',50),('B001042',50),
                 ('B001043',50),('B001044',50),('B001045',50),('B001049',50),
                 ('B001050',50),('B001051',50)],
    'LENDER-B': [('B001001',17.407),('B001027',34.712),('B001034',33.333),('B001035',50),
                 ('B001036',25),('B001037',25),('B001038',25),('B001039',25),('B001040',25),
                 ('B001041',25),('B001042',25),('B001043',25),('B001044',25),('B001045',25),
                 ('B001046',25),('B001047',25),('B001048',33.333),('B001049',25),
                 ('B001050',25),('B001051',25)],
    'LENDER-C': [('B001001',32.439),('B001008',24.3),('B001016',43.529),
                 ('B001027',19.284),('B001030',100)],
    'LENDER-D': [('B001008',40.499),('B001016',26.117),('B001019',42.172),('B001029',100)],
    'LENDER-E': [('B001005',49.915),('B001010',34.477),('B001015',29.531),
                 ('B001019',25.303),('B001020',38.983)],
    'LENDER-F': [('B001005',24.957),('B001010',22.985),('B001011',31.847),('B001020',32.486)],
    'LENDER-G': [('B001011',38.216),('B001015',41.343),('B001028',100)],
    'MI10':     [('B001001',17.407),('B001022',100),('B001023',100),('B001024',100),
                 ('B001025',100),('B001026',100),('B001027',46.004),('B001031',75),
                 ('B001032',75),('B001033',40),('B001034',33.333),('B001035',50),
                 ('B001036',25),('B001037',25),('B001038',25),('B001039',25),('B001040',25),
                 ('B001041',25),('B001042',25),('B001043',25),('B001044',25),('B001045',25),
                 ('B001046',75),('B001047',75),('B001048',66.667),('B001049',25),
                 ('B001050',25),('B001051',25)],
}

loan_lender_map = {}
for lacct, entries in portfolio_data.items():
    for loan, pct in entries:
        if pct == 0: continue
        loan_lender_map.setdefault(loan, {})[lacct] = pct

def primary_lender(acct):
    lmap = loan_lender_map.get(acct, {})
    if not lmap: return 'Unassigned'
    top = max(lmap, key=lmap.get)
    return lender_names.get(top, top)

BUCKETS = [
    ('Current',      0,   0),
    ('1-30 Days',    1,  30),
    ('31-60 Days',  31,  60),
    ('61-90 Days',  61,  90),
    ('91-120 Days', 91, 120),
    ('121+ Days',  121, 99999),
]

def get_bucket(days):
    for label, lo, hi in BUCKETS:
        if lo <= days <= hi:
            return label
    return '121+ Days'

rows = []
for l in loans:
    bal = float(l.get('PrincipalBalance') or 0)
    if bal <= 0: continue
    days  = int(l.get('DaysLate') or 0)
    acct  = l.get('Account', '')
    bname = l.get('SortName','') or l.get('ByLastName','') or ''
    if isinstance(bname, dict): bname = ''
    rows.append({
        'bucket':   get_bucket(days),
        'account':  acct,
        'borrower': bname.strip(),
        'lender':   primary_lender(acct),
        'balance':  bal,
        'days':     days,
    })

# ── Print console report ─────────────────────────────────────────────────
total_bal = sum(r['balance'] for r in rows)

bucket_totals = {}
for label, _, __ in BUCKETS:
    br  = [r for r in rows if r['bucket'] == label]
    bal = sum(r['balance'] for r in br)
    bucket_totals[label] = (len(br), bal)

W = 108
print()
print('=' * W)
print('{:^{w}}'.format('TMO LOAN AGING REPORT  |  As of April 23, 2026', w=W))
print('=' * W)
print('{:<22}  {:>8}  {:>22}  {:>10}  {}'.format(
      'Aging Bucket', '# Loans', 'Portfolio Balance', '% of Total', 'Balance Bar'))
print('-' * W)
for label, _, __ in BUCKETS:
    cnt, bal = bucket_totals[label]
    pct = (bal / total_bal * 100) if total_bal else 0
    bar = chr(9608) * int(pct / 2.5)
    print('{:<22}  {:>8}  ${:>20,.2f}  {:>9.1f}%  {}'.format(label, cnt, bal, pct, bar))
print('-' * W)
print('{:<22}  {:>8}  ${:>20,.2f}  {:>9}'.format('TOTAL', len(rows), total_bal, '100.0%'))
print('=' * W)

for label, _, __ in BUCKETS:
    br = sorted([r for r in rows if r['bucket'] == label], key=lambda x: -x['balance'])
    cnt, bal = bucket_totals[label]
    icon = {'Current':'✅','1-30 Days':'🟡','31-60 Days':'🟠','61-90 Days':'🔴','91-120 Days':'🔴','121+ Days':'🚨'}.get(label,'')
    print()
    print(f'  {icon} {label}  —  {cnt} loan{"s" if cnt!=1 else ""}  |  ${bal:,.2f}')
    print('  ' + '-'*(W-4))
    if not br:
        print('  No loans in this category.')
        continue
    print('  {:<13} {:<33} {:<30} {:>10}  {:>18}'.format(
          'Loan','Primary Borrower','Primary Lender','Days Late','Principal Balance'))
    print('  ' + '-'*13 + ' ' + '-'*33 + ' ' + '-'*30 + ' ' + '-'*10 + '  ' + '-'*18)
    for r in br:
        print('  {:<13} {:<33} {:<30} {:>10}  ${:>16,.2f}'.format(
              r['account'], r['borrower'][:32], r['lender'][:29], r['days'], r['balance']))
    print('  ' + ' '*77 + '  ' + '-'*18)
    print('  {:<13} {:<33} {:<30} {:>10}  ${:>16,.2f}'.format('SUBTOTAL','','','',bal))

print()
print('  Notes: Primary Lender = highest % ownership. Zero-balance loans excluded.')
print()

# ── Build Excel ──────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter

    BUCKET_COLORS = {
        'Current':     ('00B050', 'FFFFFF'),  # green
        '1-30 Days':   ('FFFF00', '000000'),  # yellow
        '31-60 Days':  ('FFC000', '000000'),  # orange
        '61-90 Days':  ('FF0000', 'FFFFFF'),  # red
        '91-120 Days': ('C00000', 'FFFFFF'),  # dark red
        '121+ Days':   ('7030A0', 'FFFFFF'),  # purple
    }

    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'

    title_font   = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    header_font  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    normal_font  = Font(name='Calibri', size=11)
    total_font   = Font(name='Calibri', bold=True, size=11)
    title_fill   = PatternFill('solid', fgColor='1F3864')
    header_fill  = PatternFill('solid', fgColor='2E75B6')
    total_fill   = PatternFill('solid', fgColor='D6E4F0')
    thin         = Side(style='thin', color='BFBFBF')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center       = Alignment(horizontal='center', vertical='center')
    right_align  = Alignment(horizontal='right', vertical='center')
    left_align   = Alignment(horizontal='left', vertical='center')

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'TMO Loan Aging Report  —  As of April 23, 2026'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['Aging Bucket', '# Loans', 'Portfolio Balance', '% of Total', 'Avg Days Late', 'Avg Loan Size']
    ws.append([''] * 6)  # row 2 blank
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[3].height = 20

    # Data rows
    for i, (label, _, __) in enumerate(BUCKETS):
        cnt, bal = bucket_totals[label]
        pct = (bal / total_bal * 100) if total_bal else 0
        br  = [r for r in rows if r['bucket'] == label]
        avg_days = sum(r['days'] for r in br) / cnt if cnt else 0
        avg_size = bal / cnt if cnt else 0
        bg, fg = BUCKET_COLORS[label]
        row_idx = 4 + i
        row_data = [label, cnt, bal, pct/100, avg_days, avg_size]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill  = PatternFill('solid', fgColor=bg)
            c.font  = Font(name='Calibri', size=11, color=fg, bold=(col==1))
            c.border = border
            c.alignment = center if col in (1,2) else right_align
        ws.cell(row=row_idx, column=3).number_format = '$#,##0.00'
        ws.cell(row=row_idx, column=4).number_format = '0.0%'
        ws.cell(row=row_idx, column=5).number_format = '0'
        ws.cell(row=row_idx, column=6).number_format = '$#,##0.00'

    # Totals row
    total_row = 10
    total_data = ['TOTAL', len(rows), total_bal, 1.0,
                  sum(r['days'] for r in rows)/len(rows) if rows else 0,
                  total_bal/len(rows) if rows else 0]
    for col, val in enumerate(total_data, 1):
        c = ws.cell(row=total_row, column=col, value=val)
        c.font   = total_font
        c.fill   = total_fill
        c.border = border
        c.alignment = center if col in (1,2) else right_align
    ws.cell(row=total_row, column=3).number_format = '$#,##0.00'
    ws.cell(row=total_row, column=4).number_format = '0.0%'
    ws.cell(row=total_row, column=5).number_format = '0'
    ws.cell(row=total_row, column=6).number_format = '$#,##0.00'

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18

    # ── Detail sheet ─────────────────────────────────────────────────────
    wd = wb.create_sheet('Aging Detail')

    wd.merge_cells('A1:F1')
    wd['A1'] = 'TMO Loan Aging Report — Detail  |  April 23, 2026'
    wd['A1'].font = title_font
    wd['A1'].fill = title_fill
    wd['A1'].alignment = center
    wd.row_dimensions[1].height = 28

    det_headers = ['Aging Bucket', 'Loan Account', 'Primary Borrower', 'Primary Lender', 'Days Late', 'Principal Balance']
    for col, h in enumerate(det_headers, 1):
        c = wd.cell(row=2, column=col, value=h)
        c.font      = header_font
        c.fill      = header_fill
        c.alignment = center
        c.border    = border
    wd.row_dimensions[2].height = 20

    detail_row = 3
    for label, _, __ in BUCKETS:
        br = sorted([r for r in rows if r['bucket'] == label], key=lambda x: -x['balance'])
        if not br:
            continue
        bg, fg = BUCKET_COLORS[label]
        for r in br:
            row_data = [r['bucket'], r['account'], r['borrower'], r['lender'], r['days'], r['balance']]
            for col, val in enumerate(row_data, 1):
                c = wd.cell(row=detail_row, column=col, value=val)
                c.font      = Font(name='Calibri', size=10, color='000000')
                c.border    = border
                c.alignment = left_align if col in (1,2,3,4) else right_align
            # Color the bucket column
            wd.cell(row=detail_row, column=1).fill = PatternFill('solid', fgColor=bg)
            wd.cell(row=detail_row, column=1).font = Font(name='Calibri', size=10, color=fg, bold=True)
            wd.cell(row=detail_row, column=6).number_format = '$#,##0.00'
            detail_row += 1

        # Subtotal row per bucket
        _, bal = bucket_totals[label]
        c = wd.cell(row=detail_row, column=5, value='Subtotal')
        c.font = total_font
        c.fill = PatternFill('solid', fgColor='E2EFDA')
        c.border = border
        c.alignment = right_align
        c = wd.cell(row=detail_row, column=6, value=bal)
        c.font = total_font
        c.fill = PatternFill('solid', fgColor='E2EFDA')
        c.border = border
        c.alignment = right_align
        c.number_format = '$#,##0.00'
        for col in [1,2,3,4]:
            cc = wd.cell(row=detail_row, column=col)
            cc.fill = PatternFill('solid', fgColor='E2EFDA')
            cc.border = border
        detail_row += 1

    # Grand total
    c = wd.cell(row=detail_row, column=5, value='GRAND TOTAL')
    c.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F3864')
    c.border = border
    c.alignment = right_align
    c = wd.cell(row=detail_row, column=6, value=total_bal)
    c.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F3864')
    c.border = border
    c.alignment = right_align
    c.number_format = '$#,##0.00'
    for col in [1,2,3,4]:
        cc = wd.cell(row=detail_row, column=col)
        cc.fill = PatternFill('solid', fgColor='1F3864')
        cc.border = border

    wd.column_dimensions['A'].width = 16
    wd.column_dimensions['B'].width = 14
    wd.column_dimensions['C'].width = 34
    wd.column_dimensions['D'].width = 30
    wd.column_dimensions['E'].width = 12
    wd.column_dimensions['F'].width = 22

    # Freeze header row on detail sheet
    wd.freeze_panes = 'A3'

    out = os.path.join(os.path.dirname(LOANS_FILE), 'TMO_Loan_Aging_Report.xlsx')
    wb.save(out)
    print(f'Excel saved: {out}')

except ImportError:
    print('openpyxl not installed — skipping Excel export.')
